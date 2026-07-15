#!/usr/bin/env python3
# ───────────────────────────────────────────────────────────────────────────
# India extract — "SCENARIO PLANNING TOOL Master data.xlsx" → platform schema.
# THE India pipeline (supersedes ingest-india-scenario.py / the Ram workbook):
# per Sanjay (2026-07-14) India uses ONLY this master file.
#
# One sheet ('master data'), the proven roll-up layout (col C = Data Mode):
#   Variant     → full spec, no volume      → IN_catalog (the variant library)
#   Model       → volume(U), avgCO2(AM), avgMass(AN)  → IN_fleet (sales rows)
#   Brand       → total volume              → reconciliation check
#   Group/Regulatory → illustrative FY2022 compliance rows → reference only
#
# Run:  python3 scripts/ingest-india-master.py ["path/to/master.xlsx"]
# Out:  .data/india_extract.json  (consumed by apply-india-extract.py)
# ───────────────────────────────────────────────────────────────────────────
import json, sys, os
from collections import defaultdict
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.normpath(os.path.join(ROOT, "..", "SCENARIO PLANNING TOOL Master data.xlsx"))
OUT = os.path.join(ROOT, ".data", "india_extract.json")

POWERTRAIN = {
    "ICE": "ICE", "ICE SS": "ICE", "ICE CNG": "ICE",
    "MHEV": "MHEV", "HEV": "Strong Hybrid", "PHEV": "PHEV",
    "REEV": "Range-Extender Hybrid", "BEV": "BEV", "FCEV": "BEV",
}
FUEL = {
    "Gasoline": "Petrol", "Diesel": "Diesel", "CNG": "CNG",
    "Electricity": "Electric", "Hydrogen": "Hydrogen", "LPG": "LPG",
}
DRIVECYCLE = {"ARAI": "MIDC", "MIDC": "MIDC", "NEDC": "NEDC", "WLTC": "WLTC"}
FY = lambda y: f"FY {y}-{(y + 1) % 100:02d}"
ZE_PT = {"BEV", "Range-Extender Hybrid"}

def num(v): return v if isinstance(v, (int, float)) else None
def txt(v): return str(v).strip() if v is not None else None

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["master data"]
    L = openpyxl.utils.get_column_letter
    rows = []
    for r in range(4, ws.max_row + 1):
        d = {L(c): ws.cell(row=r, column=c).value for c in range(1, 47)}
        if any(v is not None for v in d.values()):
            rows.append(d)

    variants = defaultdict(list)   # (parent, model, year) → variant rows
    models, brands, regs = [], [], []
    for d in rows:
        m = d.get("C")
        if m == "Variant": variants[(d.get("E"), d.get("G"), num(d.get("A")))].append(d)
        elif m == "Model": models.append(d)
        elif m == "Brand": brands.append(d)
        elif m in ("Regulatory", "Group"): regs.append(d)

    # ── catalog: every variant spec (master cols mirror the proven layout) ────
    catalog = []
    for (parent, model, year), specs in variants.items():
        for d in specs:
            pt = POWERTRAIN.get(d.get("L"), d.get("L") or "ICE")
            is_ze = pt in ZE_PT
            co2 = num(d.get("V"))
            rec = {
                "market": txt(d.get("B")) or "IN",
                "parent": txt(parent),
                "brand": txt(d.get("F")),
                "model": txt(model),
                "variant": txt(d.get("H")),
                "variantId": txt(d.get("I")),
                "bodyStyle": txt(d.get("J")),
                "segment": txt(d.get("K")),
                "powertrain": pt,
                "engineCC": round(num(d.get("M")) * 1000) if num(d.get("M")) else None,
                "fuel": FUEL.get(d.get("N"), txt(d.get("N"))),
                "powerKW": num(d.get("O")),
                "ftCode": txt(d.get("P")),
                "gearbox": txt(d.get("Q")),
                "driveline": txt(d.get("R")),
                "battery": num(d.get("S")),
                "kerbMass": num(d.get("T")),
                "co2": 0.0 if is_ze else co2,
                "fuelKmpl": num(d.get("W")),
                "fuelMpg": num(d.get("X")),
                "fuelL100": num(d.get("Y")),
                "footprint": num(d.get("Z")),
                "energy": num(d.get("AA")),
                "range": num(d.get("AB")),
                "rangeAlt": num(d.get("AC")),
                "otrPrice": num(d.get("AD")),
                "refMass": num(d.get("AE")),
                "testMass": num(d.get("AF")),
                "tax": num(d.get("AG")),
                "vclass": txt(d.get("AH")),
                "driveCycle": DRIVECYCLE.get(d.get("AI"), txt(d.get("AI"))),
                "lengthMm": num(d.get("AJ")),
                "widthMm": num(d.get("AK")),
                "heightMm": num(d.get("AL")),
                "year": num(d.get("A")),
            }
            catalog.append({k: v for k, v in rec.items() if v is not None})

    # ── fleet: model rows carry the sales; same fixes as the proven pipeline ──
    fleet, fixes = [], {"dropped_no_sales": [], "mass_backfilled": []}
    for d in models:
        parent, model, year = txt(d.get("E")), txt(d.get("G")), num(d.get("A"))
        vol = num(d.get("U"))
        if not vol or vol <= 0:
            fixes["dropped_no_sales"].append(f"{year} {parent[:24]} · {model}")
            continue
        avg_co2, avg_mass = num(d.get("AM")), num(d.get("AN"))
        specs = variants.get((parent, model, year), []) or variants.get((parent, model, (year or 0) + 1), [])
        pts = [POWERTRAIN.get(s.get("L"), s.get("L")) for s in specs if s.get("L")]
        modal_pt = max(set(pts), key=pts.count) if pts else "ICE"
        fuels = [FUEL.get(s.get("N"), s.get("N")) for s in specs if s.get("N")]
        modal_fuel = max(set(fuels), key=fuels.count) if fuels else "Petrol"
        is_ze = (avg_co2 == 0) or modal_pt in ZE_PT
        if avg_mass is None:
            kerbs = [num(s.get("T")) for s in specs if num(s.get("T"))]
            if kerbs:
                avg_mass = round(sum(kerbs) / len(kerbs), 1)
                fixes["mass_backfilled"].append(f"{year} {parent[:20]} · {model} → {avg_mass} kg")
        rec = {
            "parent": parent, "pool": parent,
            "brand": txt(specs[0].get("F")) if specs else model,
            "make": txt(specs[0].get("F")) if specs else model,
            "model": model, "year": year, "fyLabel": FY(year),
            "powertrain": modal_pt if specs else ("BEV" if is_ze else "ICE"),
            "fuel": "Electric" if (modal_pt in ZE_PT or is_ze) else modal_fuel,
            "co2": 0.0 if is_ze else (avg_co2 if avg_co2 is not None else 0.0),
            "mass": avg_mass,
            "sales": int(vol),
            "segment": txt(d.get("K")),
            "bodyStyle": next((txt(s.get("J")) for s in specs if s.get("J")), None),
            "footprint": next((num(s.get("Z")) for s in specs if num(s.get("Z"))), None),
            # homologation cycle joined from the model's variant specs (uniform
            # in the master: ARAI→MIDC) — powers the Drive cycle column and
            # squares with the CAFE II·MIDC / CAFE III·WLTP era bands.
            "driveCycle": next((DRIVECYCLE.get(s.get("AI"), txt(s.get("AI"))) for s in specs if s.get("AI")), None),
            "vclass": "Passenger car", "scenario": "Base",
        }
        fleet.append({k: v for k, v in rec.items() if v is not None})

    reg_rows = [{"parent": txt(g.get("E")), "targetYear": num(g.get("A")), "P_gpkm": num(g.get("AO")),
                 "T_gpkm": num(g.get("AQ")), "credit": num(g.get("AS")), "compliant": txt(g.get("AT")), "mode": txt(g.get("C"))} for g in regs]

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump({
        "meta": {"source": "SCENARIO PLANNING TOOL Master data.xlsx", "market": "IN",
                 "note": "master file is the ONLY India source (2026-07-14); regulatory/group rows are illustrative"},
        "IN_fleet": fleet, "IN_catalog": catalog, "IN_regulatory_reference": reg_rows,
    }, open(OUT, "w"), indent=1)

    # ── reconciliation ────────────────────────────────────────────────────────
    print(f"→ wrote {OUT}")
    print(f"IN_catalog: {len(catalog)} variant specs · IN_fleet: {len(fleet)} model rows")
    got = defaultdict(int)
    for v in fleet: got[(v["parent"], v["year"])] += v["sales"]
    ok = True
    for b in brands:
        key = (txt(b.get("E")), num(b.get("A")))
        expect = int(num(b.get("U")) or 0)
        g = got.get(key, 0)
        flag = "ok" if g == expect else "MISMATCH"
        ok &= flag == "ok"
        print(f"   {key[1]} {key[0][:42]:42} extracted={g:7d} workbook={expect:7d} {flag}")
    print(f"   → reconciliation {'PASSED' if ok else 'FAILED'}")
    print(f"fixes: dropped={fixes['dropped_no_sales']} · mass_backfilled={fixes['mass_backfilled']}")
    conv_missing = sum(1 for v in catalog if v["powertrain"] != "BEV" and "co2" not in v)
    print(f"catalog conventional rows missing CO₂: {conv_missing}")
    if not ok: sys.exit(1)

if __name__ == "__main__":
    main()
