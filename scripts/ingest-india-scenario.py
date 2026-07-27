#!/usr/bin/env python3
# ───────────────────────────────────────────────────────────────────────────
# India extract — "SCENARIO PLANNING TOOL UPDATE" workbook → platform schema
#
# The EU market has a live open API (see ingest-open.mjs); IN/AU/UK are
# committed workbook extracts. This is the IN extractor. It reads the Ram
# scenario workbook and emits India data in the engine's Vehicle shape:
#
#   • IN_fleet   — model-level rows that carry SALES (the compliance fleet).
#                  Source: the VIJAY sheet's `Model` roll-up rows, which are
#                  the only place a volume is recorded. One Vehicle per
#                  parent·model·year. This is what the engine weights & rolls up.
#   • IN_catalog — the full 649-variant spec library (no sales). Source: the
#                  DATA sheet. Feeds the variant picker / scenario "add variant".
#
# Run:  python3 scripts/ingest-india-scenario.py "../SCENARIO PLANNING TOOL UPDATE Ram.xlsx"
# Out:  .data/india_extract.json   (+ a validation report to stdout)
# ───────────────────────────────────────────────────────────────────────────
import json, sys, os
from collections import defaultdict
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.normpath(os.path.join(ROOT, "..", "SCENARIO PLANNING TOOL UPDATE Ram.xlsx"))
OUT = os.path.join(ROOT, ".data", "india_extract.json")

# ── value-vocabulary translation → what the India rule pack expects ──────────
# Rule pack super-credit keys: BEV, 'Range-Extender Hybrid', PHEV,
# 'Strong Hybrid Flex Fuel', 'Strong Hybrid', 'Flex Fuel Ethanol'.
POWERTRAIN = {
    "ICE": "ICE", "ICE SS": "ICE",       # start-stop is still a conventional ICE
    "ICE CNG": "ICE",                     # mislabelled in DATA (15 Tata CNG rows); fuel col already = CNG
    "MHEV": "MHEV",                        # mild hybrid — no CAFE III super-credit
    "HEV": "Strong Hybrid",               # full/strong hybrid — ×2
    "PHEV": "PHEV",                        # ×2.5
    "REEV": "Range-Extender Hybrid",      # ×3
    "BEV": "BEV",                          # ×3, zero-emission
    "FCEV": "BEV",                         # fuel-cell → zero-tailpipe, treat as ZE
}
FUEL = {
    "Gasoline": "Petrol", "Diesel": "Diesel", "CNG": "CNG",
    "Electricity": "Electric", "Hydrogen": "Hydrogen", "LPG": "LPG",
}
# 'ARAI' is the homologation agency, not a drive cycle — the Indian cycle it runs
# is MIDC. Normalise so the driveCycle field stays a valid cycle vocabulary.
DRIVECYCLE = {"ARAI": "MIDC", "MIDC": "MIDC", "NEDC": "NEDC", "WLTC": "WLTC"}
FY = lambda y: f"FY {y}-{(y + 1) % 100:02d}"          # 2025 → "FY 2025-26"
ZE_PT = {"BEV", "Range-Extender Hybrid"}               # zero-tailpipe powertrains


def col(c):
    return openpyxl.utils.get_column_letter(c)


def load(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    return wb


def rowdict(ws, r):
    return {col(c): ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)}


def num(v):
    return v if isinstance(v, (int, float)) else None


def txt(v):
    # workbook cells are occasionally numeric where a label is meant (e.g. a
    # "45 kWh" variant typed as 45) — keep string-typed fields strings.
    return str(v).strip() if v is not None else None


# ── DATA sheet → variant catalog ─────────────────────────────────────────────
# Columns (row 2 header): A Year, B Market, D Scenario, E Brand, F Model,
# G Variant, H VariantCode, I BodyStyle, J Segment, K Powertrain, L EngineCap(L),
# M Fuel, N Power(kW), O FTCode, P Gearbox, Q Driveline, R Battery(kWh),
# S KerbMass, T Volume, U FuelConsumption(CO2 g/km), V/W FuelEconomy,
# X Footprint(m2), AA VehicleClass, AD Cycle.
def parse_catalog(wb):
    ws = wb["DATA"]
    out = []
    for r in range(3, ws.max_row + 1):
        d = rowdict(ws, r)
        if not d.get("E") or not d.get("F"):
            continue
        pt = POWERTRAIN.get(d.get("K"), d.get("K") or "ICE")
        co2 = num(d.get("U"))
        is_ze = pt in ZE_PT
        rec = {
            "market": txt(d.get("B")) or "IN",
            "brand": txt(d.get("E")),
            "model": txt(d.get("F")),
            "variant": txt(d.get("G")),
            "variantId": txt(d.get("H")),
            "bodyStyle": txt(d.get("I")),
            "segment": txt(d.get("J")),
            "powertrain": pt,
            "engineCC": round(num(d.get("L")) * 1000) if num(d.get("L")) else None,
            "fuel": FUEL.get(d.get("M"), d.get("M")),
            "powerKW": num(d.get("N")),
            "gearbox": txt(d.get("P")),
            "driveline": txt(d.get("Q")),
            "battery": num(d.get("R")),
            "kerbMass": num(d.get("S")),
            "co2": 0.0 if is_ze else co2,
            "footprint": num(d.get("X")),
            "vclass": txt(d.get("AA")),
            "driveCycle": DRIVECYCLE.get(d.get("AD"), txt(d.get("AD"))),
            "year": num(d.get("A")),
        }
        out.append({k: v for k, v in rec.items() if v is not None})
    return out


# ── VIJAY sheet → sales fleet (model rows) + per-model spec (variant rows) ────
# Data Mode (col C): Variant | Model | Brand | Regulatory.
#   Variant rows: full spec, NO volume.   Model rows: volume(U), avgCO2(AM),
#   avgWeightedMass(AN).   Brand rows: total volume.   Regulatory rows: the
#   (illustrative) compliance calc AO..AT.
def parse_fleet(wb):
    ws = wb["VIJAY "]
    variants = defaultdict(list)     # (parent, model, year) → [variant spec rows]
    models, brands, regs = [], [], []
    for r in range(4, ws.max_row + 1):
        d = rowdict(ws, r)
        mode = d.get("C")
        if mode == "Variant":
            variants[(d.get("E"), d.get("G"), num(d.get("A")))].append(d)
        elif mode == "Model":
            models.append(d)
        elif mode == "Brand":
            brands.append(d)
        elif mode == "Regulatory":
            regs.append(d)

    fleet = []
    fixes = {"dropped_no_sales": [], "mass_backfilled": [], "co2_backfilled": []}
    for d in models:
        parent, model, year = d.get("E"), d.get("G"), num(d.get("A"))
        # Sales Volume column moved between master versions: BH ("Salse Volume")
        # in the 2026-07 layout, U in the earlier one. Try both.
        vol = num(d.get("BH")) or num(d.get("U"))
        # A model row with no positive volume cannot affect a sales-weighted
        # compliance average — keep it only in the catalog, not the fleet.
        if not vol or vol <= 0:
            fixes["dropped_no_sales"].append(f"{year} {(parent or '?')[:24]} · {model}")
            continue
        avg_co2 = num(d.get("AM"))
        avg_mass = num(d.get("AN"))
        specs = variants.get((parent, model, year), [])
        if specs:
            pts = [POWERTRAIN.get(s.get("L"), s.get("L")) for s in specs if s.get("L")]
            modal_pt = max(set(pts), key=pts.count) if pts else "ICE"
            fuels = [FUEL.get(s.get("N"), s.get("N")) for s in specs if s.get("N")]
            modal_fuel = max(set(fuels), key=fuels.count) if fuels else "Petrol"
            foot = next((num(s.get("Z")) for s in specs if num(s.get("Z"))), None)
            body = next((s.get("J") for s in specs if s.get("J")), None)
            kerbs = [num(s.get("T")) for s in specs if num(s.get("T"))]      # VIJAY T = Kerb Weight
            co2s = [num(s.get("V")) for s in specs if num(s.get("V"))]       # VIJAY V = CO₂ g/km
        else:
            modal_pt, modal_fuel, foot, body, kerbs, co2s = "ICE", "Petrol", None, None, [], []
        is_ze = (avg_co2 == 0) or modal_pt in ZE_PT
        # backfill mass from variant kerb weights when the model roll-up omits it
        if avg_mass is None and kerbs:
            avg_mass = round(sum(kerbs) / len(kerbs), 1)
            fixes["mass_backfilled"].append(f"{year} {parent[:20]} · {model} → {avg_mass} kg")
        # backfill CO₂ for a conventional model missing its average
        if avg_co2 is None and not is_ze and co2s:
            avg_co2 = round(sum(co2s) / len(co2s), 2)
            fixes["co2_backfilled"].append(f"{year} {parent[:20]} · {model} → {avg_co2} g/km")
        powertrain = modal_pt if specs else ("BEV" if is_ze else "ICE")
        rec = {
            "parent": txt(parent),
            "pool": txt(parent),                  # CAFE is per-manufacturer
            "brand": txt(specs[0].get("F") if specs else model),
            "make": txt(specs[0].get("F") if specs else model),
            "model": txt(model),
            "year": year,
            "fyLabel": FY(year),
            "powertrain": powertrain,
            "fuel": "Electric" if powertrain in ZE_PT else modal_fuel,
            "co2": 0.0 if is_ze else (avg_co2 if avg_co2 is not None else 0.0),
            "mass": avg_mass,
            "sales": int(vol),
            "segment": txt(d.get("K")),
            "bodyStyle": txt(body),
            "footprint": foot,
            "cnf": 0,
            "vclass": "Passenger car",
            "scenario": "Base",
        }
        fleet.append({k: v for k, v in rec.items() if v is not None})

    brand_totals = {(b.get("E"), num(b.get("A"))): (num(b.get("BH")) or num(b.get("U"))) for b in brands}
    reg_rows = [{
        "parent": g.get("E"), "targetYear": num(g.get("A")),
        "P_gpkm": num(g.get("AO")), "CAFCS_l100": num(g.get("AP")),
        "T_gpkm": num(g.get("AQ")), "ACAFC_l100": num(g.get("AR")),
        "credit": num(g.get("AS")), "compliant": g.get("AT"),
    } for g in regs]
    return fleet, brand_totals, reg_rows, fixes


# ── CO₂ backfill for catalog variants the source left blank ──────────────────
# 95/649 DATA rows (Mahindra/Tata/Toyota SUV "variant group" rows) have neither
# CO₂ nor fuel economy. All have kerb mass. Estimate tiered:
#   1. mean CO₂ of same-model+fuel siblings (mass-adjusted by the fuel slope)
#   2. same brand+segment+fuel siblings, mass-adjusted
#   3. per-fuel linear fit CO₂ = a·kerb + b over all complete ICE-family rows
# Every estimated value is flagged co2Estimated=True so consumers can badge it.
ICE_FAM = {"ICE", "Strong Hybrid", "MHEV"}  # post-normalisation powertrains


def _fit(pts):
    n = len(pts)
    mx = sum(p[0] for p in pts) / n
    my = sum(p[1] for p in pts) / n
    sxx = sum((x - mx) ** 2 for x, _ in pts) or 1e-9
    a = sum((x - mx) * (y - my) for x, y in pts) / sxx
    return a, my - a * mx


def backfill_co2(catalog):
    complete = [v for v in catalog if v.get("co2") not in (None, "") and v.get("kerbMass") and v["powertrain"] != "BEV"]
    slopes, fits = {}, {}
    for fuel in {v["fuel"] for v in complete}:
        pts = [(v["kerbMass"], v["co2"]) for v in complete if v["fuel"] == fuel]
        if len(pts) >= 8:
            a, b = _fit(pts)
            slopes[fuel], fits[fuel] = a, (a, b)
    filled = {"model": 0, "brand-segment": 0, "regression": 0}
    for v in catalog:
        if v.get("co2") is not None or v["powertrain"] == "BEV" or v.get("fuel") == "Electric":
            continue
        kerb, fuel = v.get("kerbMass"), v.get("fuel")
        a = slopes.get(fuel, 0.08)
        sib = [c for c in complete if c["model"] == v["model"] and c["fuel"] == fuel]
        tier = "model"
        if not sib:
            sib = [c for c in complete if c["brand"] == v["brand"] and c.get("segment") == v.get("segment") and c["fuel"] == fuel]
            tier = "brand-segment"
        if sib and kerb:
            base_co2 = sum(c["co2"] for c in sib) / len(sib)
            base_kerb = sum(c["kerbMass"] for c in sib) / len(sib)
            v["co2"] = round(base_co2 + a * (kerb - base_kerb), 1)
        elif fuel in fits and kerb:
            af, bf = fits[fuel]
            v["co2"] = round(af * kerb + bf, 1)
            tier = "regression"
        else:
            continue
        v["co2Estimated"] = True
        filled[tier] += 1
    return filled


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = load(xlsx)
    catalog = parse_catalog(wb)
    co2_fills = backfill_co2(catalog)
    fleet, brand_totals, regs, fixes = parse_fleet(wb)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    payload = {
        "meta": {
            "source": "SCENARIO PLANNING TOOL UPDATE Ram.xlsx",
            "market": "IN",
            "fleetGranularity": "model (sales recorded at model level in VIJAY roll-up)",
            "catalogGranularity": "variant (DATA sheet spec library, no sales)",
            "regNote": "regulatory rows are illustrative worked-example figures, not engine-computed",
        },
        "IN_fleet": fleet,
        "IN_catalog": catalog,
        "IN_regulatory_reference": regs,
    }
    with open(OUT, "w") as f:
        json.dump(payload, f, indent=1)

    # ── validation report ────────────────────────────────────────────────────
    print(f"→ wrote {OUT}")
    print(f"\nIN_catalog : {len(catalog)} variant specs")
    print(f"IN_fleet   : {len(fleet)} model rows (carry sales)")
    print(f"\nBrand-total reconciliation (sum of model sales  vs  VIJAY Brand row):")
    got = defaultdict(int)
    for v in fleet:
        got[(v["parent"], v["year"])] += v["sales"]
    ok = True
    for key, expect in sorted(brand_totals.items(), key=lambda kv: (str(kv[0][0]), kv[0][1] or 0)):
        g = got.get(key, 0)
        flag = "ok" if g == (expect or 0) else "MISMATCH"
        if flag != "ok":
            ok = False
        print(f"   {key[1]}  {(key[0] or '—')[:42]:42}  extracted={g:7d}  workbook={int(expect or 0):7d}  {flag}")
    print(f"\n   → reconciliation {'PASSED' if ok else 'FAILED'}")

    # ── data-quality corrections applied ─────────────────────────────────────
    print(f"\nData-quality corrections:")
    print(f"   'ICE CNG' powertrain normalised → 'ICE'   : {sum(1 for v in catalog if v.get('fuel')=='CNG' and v['powertrain']=='ICE')} CNG variants (fuel=CNG preserved)")
    print(f"   model rows dropped (no sales)            : {fixes['dropped_no_sales'] or 'none'}")
    print(f"   mass back-filled from variant kerb wt    : {fixes['mass_backfilled'] or 'none'}")
    print(f"   CO₂ back-filled from variant rows        : {fixes['co2_backfilled'] or 'none'}")

    # ── engine-basis compliance vs the workbook's illustrative regulatory row ─
    # The engine computes the corporate average over the WHOLE fleet (EVs count
    # at 0 g/km). Where the workbook's P differs, its worked example excluded EV
    # volume — a logical error to be aware of, not a bug in this extract.
    print(f"\nCompliance parity — engine-basis P (all vehicles, EV@0) vs workbook regulatory row:")
    agg = defaultdict(lambda: {"u": 0, "uco2": 0, "umass": 0, "uco2_ice": 0, "u_ice": 0})
    for v in fleet:
        a = agg[(v["parent"], v["year"])]
        a["u"] += v["sales"]; a["uco2"] += v["sales"] * v["co2"]; a["umass"] += v["sales"] * v.get("mass", 0)
        if v["co2"] > 0:
            a["u_ice"] += v["sales"]; a["uco2_ice"] += v["sales"] * v["co2"]
    wbP = {(g["parent"], 2025 if g["parent"] == "MG Motor" else 2026): g["P_gpkm"] for g in regs}
    for key, a in sorted(agg.items()):
        if not a["u"]:
            continue
        P_all = a["uco2"] / a["u"]
        P_ice = a["uco2_ice"] / a["u_ice"] if a["u_ice"] else 0
        wp = wbP.get(key)
        note = ""
        if wp is not None:
            near_ice = abs(wp - P_ice) < 2
            note = f" | workbook P={wp:.1f} ({'≈ICE-only avg → EVs excluded in workbook' if near_ice else 'differs'})"
        print(f"   {key[1]} {key[0][:34]:34} P_all={P_all:6.2f}  avgMass={a['umass']/a['u']:6.1f}kg{note}")

    # ── catalog data-completeness (source gaps, not extract errors) ──────────
    conv_missing = [v for v in catalog if v["powertrain"] in ("ICE", "MHEV", "Strong Hybrid", "PHEV") and "co2" not in v]
    bev_bad = [v for v in catalog if v["powertrain"] == "BEV" and v.get("co2", 0) != 0]
    est = sum(1 for v in catalog if v.get("co2Estimated"))
    print(f"\nCatalog completeness:")
    print(f"   CO₂ back-filled as flagged estimates (co2Estimated)   : {est}  (by tier: {co2_fills})")
    print(f"   conventional variants STILL missing tailpipe CO₂       : {len(conv_missing)} / {len(catalog)}")
    print(f"   BEV variants with non-zero CO₂ (should be 0)            : {len(bev_bad)}")
    print(f"   driveCycle 'ARAI' normalised → 'MIDC'                   : ok (0 'ARAI' remain: {sum(1 for v in catalog if v.get('driveCycle')=='ARAI')==0})")

    print(f"\nVocab check — powertrains: {sorted(set(v['powertrain'] for v in catalog))}")
    print(f"             fuels      : {sorted(set(v['fuel'] for v in catalog))}")
    print(f"             bodyStyles : {sorted(set(v.get('bodyStyle') for v in catalog if v.get('bodyStyle')))}")
    print(f"             segments   : {sorted(set(v.get('segment') for v in catalog if v.get('segment')))}")
    print(f"             driveCycles: {sorted(set(v.get('driveCycle') for v in catalog if v.get('driveCycle')))}")


if __name__ == "__main__":
    main()
